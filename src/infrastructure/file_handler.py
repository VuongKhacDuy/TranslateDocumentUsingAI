import os
import tempfile
from pathlib import Path
import pandas as pd
from PyPDF2 import PdfReader
from docx import Document
import xlwings as xw
import openpyxl
from openpyxl_image_loader import SheetImageLoader

class FileHandler:
    def __init__(self):
        self.input_dir = Path("input")
        self.output_dir = Path("output")
        self.images_dir = self.output_dir / "images"
        self.input_dir.mkdir(exist_ok=True)
        self.output_dir.mkdir(exist_ok=True)
        self.images_dir.mkdir(exist_ok=True)

    def save_uploaded_file(self, uploaded_file):
        """Save uploaded file to input directory"""
        input_path = self.input_dir / uploaded_file.name
        with open(input_path, "wb") as f:
            f.write(uploaded_file.getbuffer())
        return str(input_path)

    def get_mime_type(self, file_path):
        """Get MIME type based on file extension"""
        ext = Path(file_path).suffix.lower()
        mime_types = {
            '.xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            '.xls': 'application/vnd.ms-excel',
            '.pdf': 'application/pdf',
            '.doc': 'application/msword',
            '.docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
            '.csv': 'text/csv'
        }
        return mime_types.get(ext, 'application/octet-stream')

    def extract_text(self, file_path):
        """Extract text from different file types"""
        ext = Path(file_path).suffix.lower()
        
        if ext in ['.xlsx', '.xls']:
            return self._extract_from_excel(file_path)
        elif ext == '.pdf':
            return self._extract_from_pdf(file_path)
        elif ext in ['.doc', '.docx']:
            return self._extract_from_word(file_path)
        elif ext == '.csv':
            return self._extract_from_csv(file_path)
        else:
            raise ValueError(f"Unsupported file type: {ext}")

    def _extract_from_excel(self, file_path):
        """Extract text and images from Excel files"""
        texts = []
        image_locations = []

        # Extract text using xlwings
        app = xw.App(visible=False)
        try:
            wb = app.books.open(file_path)
            for sheet in wb.sheets:
                used_range = sheet.used_range
                if used_range.count > 1:
                    for cell in used_range:
                        if cell.value:
                            texts.append(str(cell.value))
        finally:
            app.quit()

        # Extract images using openpyxl
        wb = openpyxl.load_workbook(file_path)
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            image_loader = SheetImageLoader(sheet)
            
            # Find cells containing images
            for row in sheet.iter_rows():
                for cell in row:
                    if image_loader.image_in(cell.coordinate):
                        try:
                            image = image_loader.get(cell.coordinate)
                            image_path = self.images_dir / f"image_{sheet_name}_{cell.coordinate}.png"
                            image.save(str(image_path))
                            image_locations.append({
                                'sheet': sheet_name,
                                'cell': cell.coordinate,
                                'path': str(image_path)
                            })
                        except Exception as e:
                            print(f"Failed to extract image from {cell.coordinate}: {str(e)}")

        return texts, image_locations

    def save_translated_content(self, input_path, output_path, translated_texts, image_locations=None):  # Removed underscore
        """Save translated content with images"""
        file_ext = Path(input_path).suffix.lower()
        
        if file_ext in ['.xlsx', '.xls']:
            app = xw.App(visible=False)
            try:
                # Copy original file to preserve formatting
                import shutil
                shutil.copy2(input_path, output_path)
                
                # Open the copied file
                wb = app.books.open(output_path)
                try:
                    text_index = 0
                    for sheet in wb.sheets:
                        used_range = sheet.used_range
                        if used_range.count > 1:
                            for cell in used_range:
                                if cell.value and isinstance(cell.value, str):
                                    if text_index < len(translated_texts):
                                        cell.value = translated_texts[text_index]
                                        text_index += 1
                    wb.save()
                finally:
                    wb.close()
            finally:
                app.quit()

            # Restore images if any
            if image_locations:
                wb = openpyxl.load_workbook(output_path)
                for img_info in image_locations:
                    sheet = wb[img_info['sheet']]
                    img = openpyxl.drawing.image.Image(img_info['path'])
                    cell = sheet[img_info['cell']]
                    img.anchor = cell.coordinate
                    sheet.add_image(img)
                wb.save(output_path)
        else:
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write('\n'.join(translated_texts))

    def _extract_from_pdf(self, file_path):
        """Extract text from PDF files"""
        texts = []
        with open(file_path, 'rb') as file:
            pdf = PdfReader(file)
            for page in pdf.pages:
                texts.append(page.extract_text())
        return texts

    def _extract_from_word(self, file_path):
        """Extract text from Word documents"""
        doc = Document(file_path)
        return [paragraph.text for paragraph in doc.paragraphs if paragraph.text]

    def _extract_from_csv(self, file_path):
        """Extract text from CSV files"""
        df = pd.read_csv(file_path)
        texts = []
        for column in df.columns:
            texts.extend(df[column].astype(str).tolist())
        return texts